"""exportar-pedidos.py — Exporta pedidos del mes a Excel

Uso:
    python exportar-pedidos.py                          # Pedidos del mes actual
    python exportar-pedidos.py --mes 2026-04            # Abril 2026
    python exportar-pedidos.py --mes all                 # Todos los pedidos
    python exportar-pedidos.py --output pedidos.xlsx     # Nombre personalizado

Requiere DATABASE_URL en .env y openpyxl instalado.
"""

import os
import sys
import argparse
from datetime import datetime, date
from pathlib import Path
from collections import defaultdict

try:
    import psycopg2
    from psycopg2.extras import DictCursor, RealDictCursor
except ImportError:
    print("psycopg2 no instalado. Ejecuta: pip install psycopg2-binary")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl no instalado. Ejecuta: pip install openpyxl")
    sys.exit(1)


STYLE_COLORS = {
    "header_fill": PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid"),
    "header_font": Font(name="Calibri", bold=True, color="1A1A1A", size=11),
    "subheader_fill": PatternFill(start_color="F5F0E8", end_color="F5F0E8", fill_type="solid"),
    "title_font": Font(name="Calibri", bold=True, color="D4AF37", size=14),
    "data_font": Font(name="Calibri", size=10),
    "total_font": Font(name="Calibri", bold=True, size=11, color="D4AF37"),
    "border": Border(
        left=Side(style="thin", color="E0D5C1"),
        right=Side(style="thin", color="E0D5C1"),
        top=Side(style="thin", color="E0D5C1"),
        bottom=Side(style="thin", color="E0D5C1"),
    ),
}

ROW_HEIGHTS = {
    1: 8,
    2: 30,
    3: 25,
}

HEADER_COLS = [
    ("N° Pedido", 18),
    ("Fecha", 14),
    ("Cliente", 22),
    ("Email", 28),
    ("Teléfono", 16),
    ("Destino", 30),
    ("Productos", 35),
    ("Items", 8),
    ("Subtotal", 14),
    ("Envío", 10),
    ("Total", 14),
    ("Estado", 14),
    ("Pago", 14),
]


def get_connection():
    url = os.getenv("DATABASE_URL")
    if not url:
        env_path = Path(__file__).parent.parent / ".env"
        if env_path.exists():
            with open(env_path) as f:
                for line in f:
                    line = line.strip()
                    if line.startswith("DATABASE_URL="):
                        url = line.split("=", 1)[1].strip().strip('"').strip("'")
                        break
    if not url:
        print("ERROR: DATABASE_URL no encontrada en .env")
        sys.exit(1)
    return psycopg2.connect(url, sslmode="require")


def fetch_orders(conn, month_filter=None):
    where = ""
    params = []

    if month_filter and month_filter != "all":
        try:
            dt = datetime.strptime(month_filter, "%Y-%m")
            where = "WHERE o.fecha >= %s AND o.fecha < %s"
            if month_filter == "all":
                pass
            else:
                if month_filter == "all":
                    pass
                else:
                    next_month = dt.replace(day=28) + __import__("datetime").timedelta(days=4)
                    params = [dt.isoformat(), next_month.replace(day=1).isoformat()]
        except ValueError:
            pass

    if not params:
        now = datetime.now()
        first = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        next_m = first.replace(day=28) + __import__("datetime").timedelta(days=4)
        params = [first.isoformat(), next_m.replace(day=1).isoformat()]
        where = "WHERE o.fecha >= %s AND o.fecha < %s"

    query = f"""
        SELECT
            o.id, o.numero_pedido, o.fecha, o.subtotal, o.envio, o.total,
            o.estado, o.metodo_pago, o.notas,
            u.name as cliente_nombre, u.email as cliente_email,
            a.destinatario, a.direccion, a.ciudad, a.departamento, a.telefono,
            oi.product_id, oi.nombre_producto, oi.cantidad, oi.precio_unitario, oi.subtotal as item_subtotal
        FROM orders o
        JOIN users u ON u.id = o.user_id
        LEFT JOIN addresses a ON a.id = o.address_id
        LEFT JOIN order_items oi ON oi.order_id = o.id
        {where}
        ORDER BY o.fecha DESC, o.id DESC
    """

    with conn.cursor(cursor_factory=DictCursor) as cur:
        cur.execute(query, params)
        rows = cur.fetchall()

    orders = defaultdict(lambda: {
        "items": [], "productos_str": [], "total_items": 0
    })

    for row in rows:
        pid = row["id"]
        orders[pid].update({
            "numero_pedido": row["numero_pedido"],
            "fecha": row["fecha"],
            "subtotal": float(row["subtotal"] or 0),
            "envio": float(row["envio"] or 0),
            "total": float(row["total"] or 0),
            "estado": row["estado"],
            "metodo_pago": row["metodo_pago"],
            "cliente_nombre": row["cliente_nombre"],
            "cliente_email": row["cliente_email"],
            "destinatario": row["destinatario"],
            "direccion": row["direccion"],
            "ciudad": row["ciudad"],
            "departamento": row["departamento"],
            "telefono": row["telefono"],
        })
        if row["nombre_producto"]:
            orders[pid]["items"].append(dict(row))
            orders[pid]["productos_str"].append(f"{row['nombre_producto']} x{row['cantidad']}")
            orders[pid]["total_items"] += row["cantidad"]

    return list(orders.values())


def build_excel(orders, output_path, month_label):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos"

    col_widths = {i + 1: w for i, (_, w) in enumerate(HEADER_COLS)}

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADER_COLS))
    cell = ws.cell(row=1, column=1, value=f"Victorem - Reporte de Pedidos ({month_label})")
    cell.font = STYLE_COLORS["title_font"]
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = ROW_HEIGHTS[1]
    ws.row_dimensions[2].height = ROW_HEIGHTS[2]

    for col_idx, (name, width) in enumerate(HEADER_COLS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=name)
        cell.font = STYLE_COLORS["header_font"]
        cell.fill = STYLE_COLORS["header_fill"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = STYLE_COLORS["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    row_num = 3
    total_subtotal = 0
    total_envio = 0
    total_general = 0

    for order in orders:
        destino = f"{order.get('direccion', '')}, {order.get('ciudad', '')}, {order.get('departamento', '')}".strip(", ")
        fecha_str = order["fecha"].strftime("%d/%m/%Y %H:%M") if isinstance(order["fecha"], datetime) else str(order["fecha"] or "")
        productos = ", ".join(order["productos_str"])
        estado_label = {
            "pendiente": "Pendiente", "pagado": "Pagado",
            "confirmado": "Confirmado", "enviado": "Enviado",
            "entregado": "Entregado", "cancelado": "Cancelado",
        }.get(order["estado"], order["estado"])

        pago_label = {"nequi": "Nequi", "transferencia": "Transferencia", "wompi": "Wompi"}.get(
            order["metodo_pago"], order["metodo_pago"] or "N/A"
        )

        subtotal = order["subtotal"]
        envio = order["envio"]
        total = order["total"]
        total_subtotal += subtotal
        total_envio += envio
        total_general += total

        vals = [
            order["numero_pedido"], fecha_str,
            order.get("destinatario") or order["cliente_nombre"],
            order["cliente_email"], order.get("telefono", ""),
            destino, productos, order["total_items"],
            subtotal, envio, total, estado_label, pago_label,
        ]

        ws.row_dimensions[row_num].height = max(20, 15 * (len(productos) // 50 + 1))

        for col_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = STYLE_COLORS["data_font"]
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = STYLE_COLORS["border"]
            if col_idx in (9, 10, 11):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")

        row_num += 1

    row_num += 1
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
    cell = ws.cell(row=row_num, column=1, value=f"Total: {len(orders)} pedidos")
    cell.font = STYLE_COLORS["total_font"]
    cell.alignment = Alignment(horizontal="right", vertical="center")

    for col_idx, val in [(9, total_subtotal), (10, total_envio), (11, total_general)]:
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.font = STYLE_COLORS["total_font"]
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = STYLE_COLORS["border"]

    ws.sheet_properties.tabColor = "D4AF37"
    ws.freeze_panes = "A3"

    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Exporta pedidos a Excel")
    parser.add_argument("--mes", "-m", default=None,
                        help="Mes en formato YYYY-MM (ej: 2026-04). Default: mes actual. Usar 'all' para todos.")
    parser.add_argument("--output", "-o", default=None,
                        help="Nombre del archivo Excel de salida")
    args = parser.parse_args()

    conn = get_connection()

    if args.mes == "all":
        month_label = "Todos los pedidos"
    elif args.mes:
        try:
            dt = datetime.strptime(args.mes, "%Y-%m")
            month_label = dt.strftime("%B %Y").capitalize()
        except ValueError:
            print(f"Formato invalido: '{args.mes}'. Use YYYY-MM o 'all'")
            sys.exit(1)
    else:
        dt = datetime.now()
        month_label = dt.strftime("%B %Y").capitalize()

    orders = fetch_orders(conn, month_filter=args.mes or "current")
    conn.close()

    if not orders:
        print(f"No hay pedidos para {month_label}")
        return

    output = args.output or f"pedidos-{month_label.lower().replace(' ', '-')}.xlsx"
    build_excel(orders, output, month_label)

    total = sum(o["total"] for o in orders)
    print(f"\nExportados {len(orders)} pedidos a '{output}'")
    print(f"Total ventas: ${total:,.0f}")


if __name__ == "__main__":
    main()
